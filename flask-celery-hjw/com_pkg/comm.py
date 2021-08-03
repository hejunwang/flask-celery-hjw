#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@Introduce :
@File      : comm.py
@Time      : 2021/4/9 16:26
@Author    : toby
"""



import configparser,os
from db.mydb import *

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


class Comm():
    def _readcfg(self):
        '''
        读取配置文件
        '''
        db = MyDB(host='120.25.239.8', port=18001, user='autochargeadmin', passwd='autochargeadmin')
        print(db)

        return db

    def _ret_db(self,file_path,tag):
        '''
        读取数据库配置文件
        :param file_path: 文件路径
        :param tag: lab标签
        :return:
        '''
        print('-----读取数据库配置文件---------')
        conf = configparser.ConfigParser()

        conf.read(file_path, encoding='utf-8')
        print(conf.sections())

        sections = conf.sections()  # 获取配置文件中所有sections，sections是列表
        print(sections)

        host = conf.get(tag, 'host')
        port = conf.get(tag, 'port')
        user = conf.get(tag, 'user')
        passwd = conf.get(tag, 'passwd')
        print('Read config file !')
        print(host, port, user, passwd)

        db = MyDB(host=host, port=int(port), user=user, passwd=passwd)
        print(db)
        print('-----读取数据库配置文件完成,返回db---------')
        return db

    def _write_excel(self,file_name,content):
        '''
        写入到excel表格中
        :param file_name: 文件名称
        :param content:  文件内容
        :return:
        '''
        wb = Workbook()
        ws =wb.active
        print("准备写入到excel表格中。。。 ")
        try:
            if type(content) == list or tuple:
                print('list or tuple')
                for row in content:
                    ws.append(row)
            else:
                print(' not list ')
                for row in list(content):
                    ws.append(row)

        except Exception as  e:
            print('写入excel表格失败')


        # 存放位置
        wb.save("./static/sql_report/{}.xlsx".format(file_name))
        print('write_over !')

    def _read_excel(self,filename,sheet_name,rows=None,columns = None):
        '''
        读取excel中某一行或者是全部所有的数据
        :param filename:
        :param sheet_name:
        :param rows:
        :param columns:
        :return:
        '''
        wb = load_workbook(filename)
        # wb = load_workbook(filename=ExcelFullName)
        # 获取当前活跃的worksheet,默认就是第一个worksheet
        sheets = wb.get_sheet_names()
        # # 第一个表格的名称
        # sheet_first = sheets[0]
        try:
            if sheet_name not in sheets:
                print('ERROR!!! \n 输入表名错误,输入的表名是 {}'.format(sheet_name))
                import  sys
                sys.exit()
        except Exception as e:
            print('输入的Sheet name 有错')

        # # 获取特定的worksheet
        ws = wb.get_sheet_by_name(sheet_name)

        # 获取表格所有行和列，两者都是可迭代的
        # rows = ws.rows
        # print('rows-->{}'.format(rows))

        # 判断传入的输入是否是None
        if columns==None:
            _column = ws.max_column
        else:
            _column = columns
        if rows==None:
            _row = ws.max_row
        else:
            _row = rows
        # 获取某行所有值
        # columns = ws.columns
        print('传入的行数:{},列数 :{}'.format(_row,_column))
        print('行数:{},列数 :{}'.format(ws.max_row,ws.max_column ))

        # ls1 =[]
        # print('迭代所有的行')
        # for row in ws.rows:
        #     line = [col.value for col in row]
        #     print(line)
        #     ls1.append(line)
        # print(ls1)

        print('$'*120)
        print('_row:{}'.format(_row))
        print('_column:{}'.format(_column))
        ls = []
        for i in range(2, _row):
            rowdata = []
            for j in range(2,_column):

                cellvalue = ws.cell(row=i, column=j).value

                rowdata.append(cellvalue)

            ls.append(rowdata)
        print(ls)
        print('$' * 120)
        return ls

    def _read_excel_title(self,file_name,sheet_name=None):
        '''
        读取excel参数的title
        :param file_name:
        :param sheet_name:
        :return:
        '''
        wb = load_workbook(file_name)
        sheets = wb.get_sheet_names()
        if sheet_name not in sheets:
            print('ERROR!!! \n 输入表名错误,输入的表名是 {}'.format(sheet_name))
            import  sys
            sys.exit()

        # # 获取特定的worksheet
        ws = wb.get_sheet_by_name(sheet_name)
        # 第一行
        title_value=[]

        for va in ws[1]:
            title_value.append(va.value)

        print(len(title_value))
        return  title_value


    def _find_new_file(self,dir):
        '''查找目录下最新的文件'''
        file_lists = os.listdir(dir)
        file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn))
        print('最新的文件为： ' + file_lists[-1])
        file = os.path.join(dir, file_lists[-1])
        print('完整路径：', file)
        return file,file_lists[-1]


if __name__ == '__main__':
    Comm()._ret_db('../config/cfg.ini','mysql')
    # Comm()._readcfg()
    # Comm()._write_excel('test',[(1,23,3),(1,2,3)])
    ls = Comm()._read_excel('../result/1619271898.xlsx','Sheet')
    print(ls)
    pass


